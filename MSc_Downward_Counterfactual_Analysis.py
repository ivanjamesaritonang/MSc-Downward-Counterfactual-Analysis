"""
Research Project - James
Downward Counterfactual Analysis of the 2023 Turkiye-Syria Earthquakes

Building step-by-step: starting with the Pre-Code uniform scenario x 04:17 timing.
"""

import openpyxl
import numpy as np
from scipy.stats import norm
import os
import glob

# --- Locate the data file relative to THIS script (not the current working
#     directory), by checking each .xlsx candidate for the required sheet,
#     rather than assuming an exact filename. ---
REQUIRED_SHEET = 'Fragility PGA'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
all_xlsx = glob.glob(os.path.join(SCRIPT_DIR, '*.xlsx'))

DATA_FILE = None
checked = []
for path in all_xlsx:
    try:
        wb_check = openpyxl.load_workbook(path, read_only=True)
        checked.append((os.path.basename(path), list(wb_check.sheetnames)))
        if REQUIRED_SHEET in wb_check.sheetnames:
            DATA_FILE = path
            wb_check.close()
            break
        wb_check.close()
    except Exception as e:
        checked.append((os.path.basename(path), f'(could not open: {e})'))

if DATA_FILE is None:
    msg = f"Could not find any .xlsx in {SCRIPT_DIR} containing a '{REQUIRED_SHEET}' sheet.\n"
    msg += "xlsx files checked:\n"
    for name, sheets in checked:
        msg += f"  - {name}: {sheets}\n"
    raise FileNotFoundError(msg)

print(f"Using data file: {DATA_FILE}")

wb = openpyxl.load_workbook(DATA_FILE, data_only=True)

# ============================================================
# STEP 1: HAZUS fragility parameters for the Pre-Code scenario
# ============================================================

frag = {}
ws = wb['Fragility PGA']
needed_types = ['C3L', 'C3M', 'URML', 'URMM', 'W1', 'S1L', 'S1M', 'S1H']
for row in ws.iter_rows(min_row=2, values_only=True):
    design_level, btype = row[0], row[1]
    if design_level == 'Pre-Code' and btype in needed_types:
        frag[btype] = (row[9], row[10])  # (median, beta)

# ============================================================
# STEP 2: Collapse Rate and Casualty Severity 4 parameters
# ============================================================

collapse_rate = {}
ws_cr = wb['Collapse Rate']
for row in ws_cr.iter_rows(min_row=2, values_only=True):
    btype = row[0]
    if btype in needed_types:
        collapse_rate[btype] = row[3]

fr_severity4 = {}
ws_cs = wb['Casualty Severity 4']
for row in ws_cs.iter_rows(min_row=2, values_only=True):
    btype = row[0]
    if btype in needed_types:
        fr_severity4[btype] = row[3]

# ============================================================
# STEP 3: Death fraction per building type
# ============================================================

def death_fraction(btype, pga):
    median, beta = frag[btype]
    p_complete = norm.cdf(np.log(pga / median) / beta)
    return p_complete * collapse_rate[btype] * fr_severity4[btype]

# ============================================================
# STEP 4: Load province data
# ============================================================

provinces = {}
ws_summary = wb['Summary Model Input']
for row in ws_summary.iter_rows(min_row=5, max_row=15, values_only=True):
    prov = row[0]
    provinces[prov] = {
        'pga': row[1],
        'height_low': row[5], 'height_mid': row[6], 'height_high': row[7],
        'rc_frac_raw': row[19], 'mas_frac_raw': row[20], 'timber_frac_raw': row[21],
        'steel_frac_raw': row[22], 'other_frac_raw': row[23],
        'occ_night': row[25],
    }

for p, d in provinces.items():
    known_total = d['rc_frac_raw'] + d['mas_frac_raw'] + d['timber_frac_raw'] + d['steel_frac_raw']
    d['rc_bldg_frac'] = d['rc_frac_raw'] / known_total
    d['mas_bldg_frac'] = d['mas_frac_raw'] / known_total
    d['timber_bldg_frac'] = d['timber_frac_raw'] / known_total
    d['steel_bldg_frac'] = d['steel_frac_raw'] / known_total

def redistribute_height(hfrac_low, hfrac_mid, hfrac_high):
    new_low = hfrac_low
    new_mid = hfrac_mid + hfrac_high
    return new_low, new_mid

def compute_province_deaths_precode_0417(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_night']
    total = 0.0

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    for mat_key, frac_key, btype_low, btype_mid in [
        ('rc', 'rc_bldg_frac', 'C3L', 'C3M'),
        ('mas', 'mas_bldg_frac', 'URML', 'URMM'),
    ]:
        mat_frac = prov_data[frac_key]
        total += occ * mat_frac * low_r * death_fraction(btype_low, pga)
        total += occ * mat_frac * mid_r * death_fraction(btype_mid, pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction('W1', pga)

    return total

national_total = sum(compute_province_deaths_precode_0417(d) for d in provinces.values())

# ============================================================
# STEP 5: Pre-Code x 14:00
# ============================================================

ws_summary2 = wb['Summary Model Input']
for row in ws_summary2.iter_rows(min_row=5, max_row=15, values_only=True):
    prov = row[0]
    provinces[prov]['occ_day'] = row[26]

def compute_province_deaths_precode_1400(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_day']
    total = 0.0

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    for mat_key, frac_key, btype_low, btype_mid in [
        ('rc', 'rc_bldg_frac', 'C3L', 'C3M'),
        ('mas', 'mas_bldg_frac', 'URML', 'URMM'),
    ]:
        mat_frac = prov_data[frac_key]
        total += occ * mat_frac * low_r * death_fraction(btype_low, pga)
        total += occ * mat_frac * mid_r * death_fraction(btype_mid, pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction('W1', pga)

    return total

national_total_1400 = sum(compute_province_deaths_precode_1400(d) for d in provinces.values())

# ============================================================
# STEP 6: Pre-Code x 17:00
# ============================================================

ws_summary3 = wb['Summary Model Input']
for row in ws_summary3.iter_rows(min_row=5, max_row=15, values_only=True):
    prov = row[0]
    provinces[prov]['occ_transit'] = row[27]

def compute_province_deaths_precode_1700(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_transit']
    total = 0.0

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    for mat_key, frac_key, btype_low, btype_mid in [
        ('rc', 'rc_bldg_frac', 'C3L', 'C3M'),
        ('mas', 'mas_bldg_frac', 'URML', 'URMM'),
    ]:
        mat_frac = prov_data[frac_key]
        total += occ * mat_frac * low_r * death_fraction(btype_low, pga)
        total += occ * mat_frac * mid_r * death_fraction(btype_mid, pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction('W1', pga)

    return total

national_total_1700 = sum(compute_province_deaths_precode_1700(d) for d in provinces.values())

# ============================================================
# STEP 7: Low-Code x 04:17
# ============================================================

frag_lowcode = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    design_level, btype = row[0], row[1]
    if design_level == 'Low-Code' and btype in needed_types:
        frag_lowcode[btype] = (row[9], row[10])

def death_fraction_lowcode(btype, pga):
    median, beta = frag_lowcode[btype]
    p_complete = norm.cdf(np.log(pga / median) / beta)
    return p_complete * collapse_rate[btype] * fr_severity4[btype]

def compute_province_deaths_lowcode_0417(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_night']
    total = 0.0

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    for mat_key, frac_key, btype_low, btype_mid in [
        ('rc', 'rc_bldg_frac', 'C3L', 'C3M'),
        ('mas', 'mas_bldg_frac', 'URML', 'URMM'),
    ]:
        mat_frac = prov_data[frac_key]
        total += occ * mat_frac * low_r * death_fraction_lowcode(btype_low, pga)
        total += occ * mat_frac * mid_r * death_fraction_lowcode(btype_mid, pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction_lowcode('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction_lowcode('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction_lowcode('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction_lowcode('W1', pga)

    return total

national_total_lowcode_0417 = sum(compute_province_deaths_lowcode_0417(d) for d in provinces.values())

# ============================================================
# STEP 8: Low-Code x 14:00
# ============================================================

def compute_province_deaths_lowcode_1400(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_day']
    total = 0.0

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    for mat_key, frac_key, btype_low, btype_mid in [
        ('rc', 'rc_bldg_frac', 'C3L', 'C3M'),
        ('mas', 'mas_bldg_frac', 'URML', 'URMM'),
    ]:
        mat_frac = prov_data[frac_key]
        total += occ * mat_frac * low_r * death_fraction_lowcode(btype_low, pga)
        total += occ * mat_frac * mid_r * death_fraction_lowcode(btype_mid, pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction_lowcode('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction_lowcode('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction_lowcode('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction_lowcode('W1', pga)

    return total

national_total_lowcode_1400 = sum(compute_province_deaths_lowcode_1400(d) for d in provinces.values())

# ============================================================
# STEP 9: Low-Code x 17:00
# ============================================================

def compute_province_deaths_lowcode_1700(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_transit']
    total = 0.0

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    for mat_key, frac_key, btype_low, btype_mid in [
        ('rc', 'rc_bldg_frac', 'C3L', 'C3M'),
        ('mas', 'mas_bldg_frac', 'URML', 'URMM'),
    ]:
        mat_frac = prov_data[frac_key]
        total += occ * mat_frac * low_r * death_fraction_lowcode(btype_low, pga)
        total += occ * mat_frac * mid_r * death_fraction_lowcode(btype_mid, pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction_lowcode('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction_lowcode('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction_lowcode('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction_lowcode('W1', pga)

    return total

national_total_lowcode_1700 = sum(compute_province_deaths_lowcode_1700(d) for d in provinces.values())

# ============================================================
# STEP 10: High-Code x 04:17
# ============================================================

frag_highcode = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    design_level, btype = row[0], row[1]
    if design_level == 'High-Code' and btype in ['C1L', 'C1M', 'C1H', 'W1', 'S1L', 'S1M', 'S1H']:
        frag_highcode[btype] = (row[9], row[10])
frag_highcode['URML'] = frag_lowcode['URML']
frag_highcode['URMM'] = frag_lowcode['URMM']

collapse_rate_hc = dict(collapse_rate)
fr_severity4_hc = dict(fr_severity4)
for row in wb['Collapse Rate'].iter_rows(min_row=2, values_only=True):
    if row[0] in ['C1L', 'C1M', 'C1H']:
        collapse_rate_hc[row[0]] = row[3]
for row in wb['Casualty Severity 4'].iter_rows(min_row=2, values_only=True):
    if row[0] in ['C1L', 'C1M', 'C1H']:
        fr_severity4_hc[row[0]] = row[3]

def death_fraction_highcode(btype, pga):
    median, beta = frag_highcode[btype]
    p_complete = norm.cdf(np.log(pga / median) / beta)
    return p_complete * collapse_rate_hc[btype] * fr_severity4_hc[btype]

def compute_province_deaths_highcode_0417(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_night']
    total = 0.0

    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_low']  * death_fraction_highcode('C1L', pga)
    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_mid']  * death_fraction_highcode('C1M', pga)
    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_high'] * death_fraction_highcode('C1H', pga)

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    total += occ * prov_data['mas_bldg_frac'] * low_r * death_fraction_highcode('URML', pga)
    total += occ * prov_data['mas_bldg_frac'] * mid_r * death_fraction_highcode('URMM', pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction_highcode('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction_highcode('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction_highcode('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction_highcode('W1', pga)

    return total

national_total_highcode_0417 = sum(compute_province_deaths_highcode_0417(d) for d in provinces.values())

# ============================================================
# STEP 11: High-Code x 14:00
# ============================================================

def compute_province_deaths_highcode_1400(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_day']
    total = 0.0

    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_low']  * death_fraction_highcode('C1L', pga)
    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_mid']  * death_fraction_highcode('C1M', pga)
    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_high'] * death_fraction_highcode('C1H', pga)

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    total += occ * prov_data['mas_bldg_frac'] * low_r * death_fraction_highcode('URML', pga)
    total += occ * prov_data['mas_bldg_frac'] * mid_r * death_fraction_highcode('URMM', pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction_highcode('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction_highcode('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction_highcode('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction_highcode('W1', pga)

    return total

national_total_highcode_1400 = sum(compute_province_deaths_highcode_1400(d) for d in provinces.values())

# ============================================================
# STEP 12: High-Code x 17:00
# ============================================================

def compute_province_deaths_highcode_1700(prov_data):
    pga = prov_data['pga']
    occ = prov_data['occ_transit']
    total = 0.0

    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_low']  * death_fraction_highcode('C1L', pga)
    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_mid']  * death_fraction_highcode('C1M', pga)
    total += occ * prov_data['rc_bldg_frac'] * prov_data['height_high'] * death_fraction_highcode('C1H', pga)

    low_r, mid_r = redistribute_height(prov_data['height_low'], prov_data['height_mid'], prov_data['height_high'])
    total += occ * prov_data['mas_bldg_frac'] * low_r * death_fraction_highcode('URML', pga)
    total += occ * prov_data['mas_bldg_frac'] * mid_r * death_fraction_highcode('URMM', pga)

    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_low']  * death_fraction_highcode('S1L', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_mid']  * death_fraction_highcode('S1M', pga)
    total += occ * prov_data['steel_bldg_frac'] * prov_data['height_high'] * death_fraction_highcode('S1H', pga)

    total += occ * prov_data['timber_bldg_frac'] * death_fraction_highcode('W1', pga)

    return total

national_total_highcode_1700 = sum(compute_province_deaths_highcode_1700(d) for d in provinces.values())

print("\n" + "="*60)
print("FULL 3x3 MATRIX (National Total Deaths) -- HAZUS")
print("="*60)
print(f"{'':12} {'04:17':>10} {'14:00':>10} {'17:00':>10}")
print(f"{'Pre-Code':12} {national_total:10,.1f} {national_total_1400:10,.1f} {national_total_1700:10,.1f}")
print(f"{'Low-Code':12} {national_total_lowcode_0417:10,.1f} {national_total_lowcode_1400:10,.1f} {national_total_lowcode_1700:10,.1f}")
print(f"{'High-Code':12} {national_total_highcode_0417:10,.1f} {national_total_highcode_1400:10,.1f} {national_total_highcode_1700:10,.1f}")
print(f"\nActual benchmark (Chen et al. 2025, factual scenario, mixed era -- reference only): 52,724")

# ============================================================
# STEP 13: Factual scenario (actual code-era mix x 04:17)
# ============================================================

ws_summary_era = wb['Summary Model Input']
for row in ws_summary_era.iter_rows(min_row=5, max_row=15, values_only=True):
    prov = row[0]
    provinces[prov]['precode_pct'] = row[2]
    provinces[prov]['lowcode_pct'] = row[3]
    provinces[prov]['highcode_pct'] = row[4]

def compute_province_deaths_factual_0417(prov_data):
    deaths_pre  = compute_province_deaths_precode_0417(prov_data)
    deaths_low  = compute_province_deaths_lowcode_0417(prov_data)
    deaths_high = compute_province_deaths_highcode_0417(prov_data)

    return (
        prov_data['precode_pct']  * deaths_pre +
        prov_data['lowcode_pct']  * deaths_low +
        prov_data['highcode_pct'] * deaths_high
    )

national_total_factual = sum(compute_province_deaths_factual_0417(d) for d in provinces.values())

print(f"\nNational total (FACTUAL, actual code-era mix x 04:17): {national_total_factual:,.1f}")
print(f"Actual benchmark (Chen et al. 2025 / MoEUCC):            52,724")
print(f"Model / Actual ratio: {national_total_factual/52724:.3f}")

# ============================================================
# STEP 14: Retrofit sweep (national + per-province)
# ============================================================
import pandas as pd
import matplotlib.pyplot as plt

def compute_province_deaths_retrofit_0417(prov_data, r):
    deaths_pre  = compute_province_deaths_precode_0417(prov_data)
    deaths_low  = compute_province_deaths_lowcode_0417(prov_data)
    deaths_high = compute_province_deaths_highcode_0417(prov_data)

    pre_pct  = prov_data['precode_pct']  * (1 - r)
    low_pct  = prov_data['lowcode_pct']  * (1 - r)
    high_pct = prov_data['highcode_pct'] + r * (prov_data['precode_pct'] + prov_data['lowcode_pct'])

    return pre_pct * deaths_pre + low_pct * deaths_low + high_pct * deaths_high

RETROFIT_LEVELS = np.round(np.arange(0.0, 1.0001, 0.05), 2)

retrofit_national = {}
retrofit_by_province = {r: {} for r in RETROFIT_LEVELS}
for r in RETROFIT_LEVELS:
    total = 0.0
    for prov, d in provinces.items():
        deaths_r = compute_province_deaths_retrofit_0417(d, r)
        retrofit_by_province[r][prov] = deaths_r
        total += deaths_r
    retrofit_national[r] = total

df_retrofit = pd.DataFrame(
    {prov: [retrofit_by_province[r][prov] for r in RETROFIT_LEVELS] for prov in provinces},
    index=[f"{r*100:.0f}%" for r in RETROFIT_LEVELS]
)
df_retrofit.index.name = 'Retrofit %'
df_retrofit['NATIONAL TOTAL'] = [retrofit_national[r] for r in RETROFIT_LEVELS]
df_retrofit.to_excel('retrofit_sweep_results.xlsx')
print("\nSaved: retrofit_sweep_results.xlsx")

r_pct = RETROFIT_LEVELS * 100

r_grid = np.array(RETROFIT_LEVELS)
deaths_grid = np.array([retrofit_national[r] for r in RETROFIT_LEVELS])
print("\nRetrofit % needed to reach selected reduction targets from Factual")
for target_pct_reduction in [25, 50, 75, 90]:
    target_deaths = national_total_factual * (1 - target_pct_reduction/100)
    if target_deaths < deaths_grid.min():
        print(f"  {target_pct_reduction:3d}% reduction: NOT achievable even at 100% retrofit "
              f"(100% retrofit only reaches {(1 - deaths_grid.min()/national_total_factual)*100:.1f}% reduction)")
    else:
        r_needed = np.interp(target_deaths, deaths_grid[::-1], r_grid[::-1])
        print(f"  {target_pct_reduction:3d}% reduction (target deaths={target_deaths:,.0f}): "
              f"requires retrofitting ~{r_needed*100:.0f}% of non-High-Code stock")

# ============================================================
# STEP 15: Retrofit sweep, HAZUS -- all 3 timing scenarios
# ============================================================

HAZUS_TIME_FUNCS = {
    '04:17': (compute_province_deaths_precode_0417, compute_province_deaths_lowcode_0417, compute_province_deaths_highcode_0417),
    '14:00': (compute_province_deaths_precode_1400, compute_province_deaths_lowcode_1400, compute_province_deaths_highcode_1400),
    '17:00': (compute_province_deaths_precode_1700, compute_province_deaths_lowcode_1700, compute_province_deaths_highcode_1700),
}

def compute_province_deaths_retrofit(prov_data, r, time_label):
    pre_fn, low_fn, high_fn = HAZUS_TIME_FUNCS[time_label]
    deaths_pre  = pre_fn(prov_data)
    deaths_low  = low_fn(prov_data)
    deaths_high = high_fn(prov_data)

    pre_pct  = prov_data['precode_pct']  * (1 - r)
    low_pct  = prov_data['lowcode_pct']  * (1 - r)
    high_pct = prov_data['highcode_pct'] + r * (prov_data['precode_pct'] + prov_data['lowcode_pct'])

    return pre_pct * deaths_pre + low_pct * deaths_low + high_pct * deaths_high

national_total_factual_1400 = sum(compute_province_deaths_retrofit(d, 0.0, '14:00') for d in provinces.values())
national_total_factual_1700 = sum(compute_province_deaths_retrofit(d, 0.0, '17:00') for d in provinces.values())

retrofit_national_by_time = {t: {} for t in ['04:17', '14:00', '17:00']}
for time_label in ['04:17', '14:00', '17:00']:
    for r in RETROFIT_LEVELS:
        total = sum(compute_province_deaths_retrofit(d, r, time_label) for d in provinces.values())
        retrofit_national_by_time[time_label][r] = total

df_retrofit_alltime = pd.DataFrame(
    {t: [retrofit_national_by_time[t][r] for r in RETROFIT_LEVELS] for t in ['04:17', '14:00', '17:00']},
    index=[f"{r*100:.0f}%" for r in RETROFIT_LEVELS]
)
df_retrofit_alltime.index.name = 'Retrofit %'
df_retrofit_alltime.to_excel('retrofit_sweep_hazus_all_times.xlsx')
print("\nSaved: retrofit_sweep_hazus_all_times.xlsx")

def print_retrofit_targets(retrofit_dict, time_label, model_name):
    baseline = retrofit_dict[time_label][0.0]
    r_grid_t = np.array(RETROFIT_LEVELS)
    deaths_grid_t = np.array([retrofit_dict[time_label][r] for r in RETROFIT_LEVELS])
    floor = deaths_grid_t.min()
    max_reduction_pct = (1 - floor / baseline) * 100
    print(f"\n{model_name} -- retrofit % needed to reach reduction targets, {time_label} "
          f"(baseline={baseline:,.1f}, floor={floor:,.1f}):")
    for target_pct_reduction in [25, 50, 75, 90]:
        target_deaths = baseline * (1 - target_pct_reduction/100)
        if target_deaths < floor:
            print(f"  {target_pct_reduction:3d}% reduction: NOT achievable even at 100% retrofit "
                  f"(100% retrofit only reaches {max_reduction_pct:.1f}% reduction)")
        else:
            r_needed = np.interp(target_deaths, deaths_grid_t[::-1], r_grid_t[::-1])
            print(f"  {target_pct_reduction:3d}% reduction (target deaths={target_deaths:,.0f}): "
                  f"requires retrofitting ~{r_needed*100:.0f}% of non-High-Code stock")

print("\n" + "="*70)
print("Diminishing-returns check across all 3 timings, HAZUS (Figure 12 data)")
print("="*70)
for time_label in ['04:17', '14:00', '17:00']:
    print_retrofit_targets(retrofit_national_by_time, time_label, 'HAZUS')

colors_time = {'04:17': '#2a78d6', '14:00': '#eda100', '17:00': '#3aa655'}
fig, ax = plt.subplots(figsize=(8, 5.5))
for time_label in ['04:17', '14:00', '17:00']:
    vals = [retrofit_national_by_time[time_label][r] for r in RETROFIT_LEVELS]
    ax.plot(r_pct, vals, marker='o', markersize=3, linewidth=2,
            color=colors_time[time_label], label=time_label)
ax.set_xlabel('Retrofit share of non-High-Code stock upgraded to High-Code (%)')
ax.set_ylabel('National total deaths')
ax.set_title('HAZUS retrofit sensitivity, all 3 timing scenarios')
ax.legend(title='Event time', fontsize=9)
ax.grid(alpha=0.3)
plt.tight_layout()
plt.savefig('retrofit_sweep_hazus_all_times.png', dpi=150)
plt.close()
print("Saved: retrofit_sweep_hazus_all_times.png")

sq1_factual_ratio_1400 = national_total_factual / national_total_factual_1400
sq1_factual_ratio_1700 = national_total_factual / national_total_factual_1700
sq2_factual_ratio = national_total_factual / national_total_highcode_0417

print("\nFactual-weighted SQ1 and SQ2 ratios (using actual 2023 compliance mix)")
print(f"SQ1 -- Factual 04:17 / Factual 14:00: {sq1_factual_ratio_1400:.3f}")
print(f"SQ1 -- Factual 04:17 / Factual 17:00: {sq1_factual_ratio_1700:.3f}")
print(f"SQ2 -- Factual 04:17 / Uniform High-Code 04:17: {sq2_factual_ratio:.3f}")

# ============================================================
# STEP 16: Combined SQ1 + SQ2 sensitivity chart -- HAZUS
# ============================================================

fig, ax = plt.subplots(figsize=(9.5, 6.5))
LABEL_POINTS = [0, 20, 40, 60, 80, 100]
for time_label in ['04:17', '14:00', '17:00']:
    vals_pct = [(1 - retrofit_national_by_time[time_label][r] / national_total_factual) * 100 for r in RETROFIT_LEVELS]
    ax.plot(r_pct, vals_pct, marker='o', markersize=3, linewidth=2,
            color=colors_time[time_label], label=time_label)
    for lp in LABEL_POINTS:
        r_key = round(lp / 100, 2)
        val_at_lp = (1 - retrofit_national_by_time[time_label][r_key] / national_total_factual) * 100
        ax.annotate(f'{val_at_lp:.1f}%', xy=(lp, val_at_lp), xytext=(0, 7),
                    textcoords='offset points', ha='center', fontsize=8,
                    color=colors_time[time_label], fontweight='medium')

ax.axhline(0, color='gray', linestyle=':', linewidth=0.8)
ax.annotate('Factual (04:17, r=0) = 0% reduction', xy=(2, 2), fontsize=8, color='gray')
ax.set_xticks(LABEL_POINTS)
ax.set_xlabel('Retrofit share of non-High-Code stock upgraded to High-Code (%)')
ax.set_ylabel('% reduction in deaths, relative to Factual (04:17, r=0) baseline')
ax.set_title('HAZUS: combined SQ1 (timing) + SQ2 (retrofit) sensitivity\n'
              '(vertical gap at r=0 = SQ1 effect; slope of each line = SQ2 effect)')
ax.legend(title='Event time', fontsize=9)
ax.grid(alpha=0.3)
plt.tight_layout()
plt.savefig('sensitivity_combined_hazus.png', dpi=150)
plt.close()
print("Saved: sensitivity_combined_hazus.png")

# ============================================================
# STEP 17: Export the full 3x3 matrix to Excel
# ============================================================

matrix_data = {
    '04:17': [national_total, national_total_lowcode_0417, national_total_highcode_0417],
    '14:00': [national_total_1400, national_total_lowcode_1400, national_total_highcode_1400],
    '17:00': [national_total_1700, national_total_lowcode_1700, national_total_highcode_1700],
}
df_matrix = pd.DataFrame(matrix_data, index=['Pre-Code', 'Low-Code', 'High-Code'])
df_matrix.index.name = 'SQ2 (Compliance)'

with pd.ExcelWriter('HAZUS_3x3_Matrix_Results.xlsx') as writer:
    df_matrix.to_excel(writer, sheet_name='National Totals')

    cell_functions = {
        ('Pre-Code','04:17'): compute_province_deaths_precode_0417,
        ('Pre-Code','14:00'): compute_province_deaths_precode_1400,
        ('Pre-Code','17:00'): compute_province_deaths_precode_1700,
        ('Low-Code','04:17'): compute_province_deaths_lowcode_0417,
        ('Low-Code','14:00'): compute_province_deaths_lowcode_1400,
        ('Low-Code','17:00'): compute_province_deaths_lowcode_1700,
        ('High-Code','04:17'): compute_province_deaths_highcode_0417,
        ('High-Code','14:00'): compute_province_deaths_highcode_1400,
        ('High-Code','17:00'): compute_province_deaths_highcode_1700,
    }
    province_rows = []
    for prov, d in provinces.items():
        row = {'Province': prov, 'PGA': d['pga']}
        for (era, time), fn in cell_functions.items():
            row[f'{era} x {time}'] = round(fn(d), 1)
        province_rows.append(row)
    df_province = pd.DataFrame(province_rows)
    df_province.to_excel(writer, sheet_name='Per-Province Detail', index=False)

print("Saved: HAZUS_3x3_Matrix_Results.xlsx (National Totals + Per-Province Detail sheets)")

# ============================================================
# STEP 18: HAZUS 3x3 heatmap (national totals)
#
# Visual counterpart to the GEM 3x3 heatmap (Step 30). Uses
# df_matrix, already built in Step 17 -- no new totals are
# computed here, this only visualises numbers that already exist.
# ============================================================

fig, ax = plt.subplots(figsize=(6, 4.5))
matrix_vals_hazus = df_matrix.values
im = ax.imshow(matrix_vals_hazus, cmap='Blues', aspect='auto')

ax.set_xticks(range(3))
ax.set_xticklabels(['04:17\n(night-time)', '14:00\n(day-time)', '17:00\n(transit)'])
ax.set_yticks(range(3))
ax.set_yticklabels(['Pre-Code', 'Low-Code', 'High-Code'])

for i in range(3):
    for j in range(3):
        val = matrix_vals_hazus[i, j]
        text_color = 'white' if val > matrix_vals_hazus.max() * 0.5 else 'black'
        label = f'{val:,.0f}' if val >= 1 else f'{val:.3f}'
        ax.text(j, i, label, ha='center', va='center', color=text_color, fontsize=11)

ax.set_title('HAZUS 3x3 matrix: national total deaths')
fig.colorbar(im, ax=ax, label='Deaths')
plt.tight_layout()
plt.savefig('hazus_3x3_heatmap.png', dpi=150)
plt.close()
print("Saved: hazus_3x3_heatmap.png")
print("\nHAZUS 3x3 Matrix (National Totals):")
print(df_matrix)

# ============================================================
# STEP 19: SQ1 vs SQ2 sensitivity ratio chart -- HAZUS
# ============================================================

sq1_ratios = {
    'Pre-Code': national_total / national_total_1400,
    'Low-Code': national_total_lowcode_0417 / national_total_lowcode_1400,
    'High-Code': national_total_highcode_0417 / national_total_highcode_1400,
}
sq2_ratios = {
    '04:17': national_total / national_total_highcode_0417,
    '14:00': national_total_1400 / national_total_highcode_1400,
    '17:00': national_total_1700 / national_total_highcode_1700,
}

TIME_LABELS_DISPLAY = {'04:17': '04:17\n(night-time)', '14:00': '14:00\n(day-time)', '17:00': '17:00\n(transit)'}

fig, ax = plt.subplots(figsize=(8, 5.5))
labels = list(sq1_ratios.keys()) + [TIME_LABELS_DISPLAY[t] for t in sq2_ratios.keys()]
values = list(sq1_ratios.values()) + list(sq2_ratios.values())
colors = ['#2a78d6']*3 + ['#eda100']*3

bars = ax.bar(labels, values, color=colors)
ax.set_ylabel('Ratio (worst / best)')
ax.set_title('SQ1 (timing) vs SQ2 (compliance) sensitivity ratio, HAZUS')
for bar, val in zip(bars, values):
    ax.text(bar.get_x() + bar.get_width()/2, val + max(values)*0.02, f'{val:.2f}', ha='center', fontsize=10)
plt.xticks(rotation=0, ha='center')

ax.annotate('SQ1 ratio, computed with compliance\nHELD FIXED at each level shown below',
            xy=(1, 0), xycoords=('data', 'axes fraction'),
            xytext=(0, -60), textcoords='offset points',
            ha='center', fontsize=7.5, color='#1a4e8a')
ax.annotate('SQ2 ratio, computed with timing\nHELD FIXED at each time shown below',
            xy=(4, 0), xycoords=('data', 'axes fraction'),
            xytext=(0, -60), textcoords='offset points',
            ha='center', fontsize=7.5, color='#a06800')

ax.set_ylim(0, max(values) * 1.3)

from matplotlib.patches import Patch
ax.legend(handles=[
    Patch(color='#2a78d6', label='SQ1 ratio, per compliance level'),
    Patch(color='#eda100', label='SQ2 ratio, per timing'),
], loc='upper center', bbox_to_anchor=(0.5, -0.35), ncol=2, fontsize=8, frameon=False)

plt.subplots_adjust(bottom=0.38)
plt.savefig('sq1_sq2_sensitivity.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: sq1_sq2_sensitivity.png")

print("\nSQ1 sensitivity ratios (04:17 / 14:00), per compliance level")
for level, ratio in sq1_ratios.items():
    print(f"  {level:12}: {ratio:.3f}")
print(f"  {'Range':12}: {min(sq1_ratios.values()):.3f}-{max(sq1_ratios.values()):.3f}")

print("\nSQ2 sensitivity ratios (Pre-Code / High-Code), per timing")
for time_label, ratio in sq2_ratios.items():
    print(f"  {time_label:12}: {ratio:.3f}")
print(f"  {'Range':12}: {min(sq2_ratios.values()):.3f}-{max(sq2_ratios.values()):.3f}")

print("\nInterpretation check: which parameter dominates? (factual-weighted basis)")
if sq2_factual_ratio > sq1_factual_ratio_1400:
    print(f"  SQ2 (compliance) factual ratio ({sq2_factual_ratio:.2f}x) > SQ1 (timing) factual ratio ({sq1_factual_ratio_1400:.2f}x)")
    print("  --> Building code compliance is the more sensitive/dominant parameter.")
else:
    print(f"  SQ1 (timing) factual ratio ({sq1_factual_ratio_1400:.2f}x) >= SQ2 (compliance) factual ratio ({sq2_factual_ratio:.2f}x)")
    print("  --> Timing and compliance sensitivity are comparable in magnitude under the real 2023 building stock;")
    print("      neither dominates decisively once anchored to the actual compliance mix.")

# ============================================================
# STEP 20: Interaction effect -- SQ1 x SQ2, HAZUS
# ============================================================

M = np.array([
    [national_total, national_total_1400, national_total_1700],
    [national_total_lowcode_0417, national_total_lowcode_1400, national_total_lowcode_1700],
    [national_total_highcode_0417, national_total_highcode_1400, national_total_highcode_1700],
])
rows_lbl = ['Pre-Code', 'Low-Code', 'High-Code']
cols_lbl = ['04:17', '14:00', '17:00']
COLS_LBL_SHORT = {'04:17': '04:17 (night)', '14:00': '14:00 (day)', '17:00': '17:00 (transit)'}

logM = np.log(M)
grand_mean = logM.mean()
row_eff = logM.mean(axis=1) - grand_mean
col_eff = logM.mean(axis=0) - grand_mean

interaction = np.zeros_like(logM)
for i in range(3):
    for j in range(3):
        interaction[i, j] = logM[i, j] - grand_mean - row_eff[i] - col_eff[j]

print("\nInteraction residuals (% deviation from additive model), HAZUS:")
for i, r in enumerate(rows_lbl):
    pct = [f"{(np.exp(interaction[i,j]) - 1) * 100:+.1f}%" for j in range(3)]
    print(f"  {r:12} {pct}")

idx = np.unravel_index(np.argmax(np.abs(interaction)), interaction.shape)
max_dev_pct = (np.exp(interaction[idx]) - 1) * 100
print(f"\nLargest interaction: {rows_lbl[idx[0]]} x {cols_lbl[idx[1]]} -- deviates {max_dev_pct:+.1f}% from the additive prediction")

hazus_max_interaction_cell = (rows_lbl[idx[0]], cols_lbl[idx[1]])
hazus_max_interaction_pct = max_dev_pct

# ============================================================
# STEP 21: Tornado charts (sorted horizontal bar)
# ============================================================

interaction_list = []
for i, r in enumerate(rows_lbl):
    for j, c in enumerate(cols_lbl):
        pct = (np.exp(interaction[i, j]) - 1) * 100
        interaction_list.append((f'{r} x {COLS_LBL_SHORT[c]}', pct))
interaction_list.sort(key=lambda t: abs(t[1]))

fig, ax = plt.subplots(figsize=(7.5, 5))
labels = [t[0] for t in interaction_list]
vals = [t[1] for t in interaction_list]
colors = ['#e34948' if v < 0 else '#2a78d6' for v in vals]
ax.barh(labels, vals, color=colors)
ax.axvline(0, color='black', linewidth=0.8)
ax.set_xlabel('Deviation from additive model (%)')
ax.set_title('Interaction residual by cell, ranked by |deviation|, HAZUS')
max_abs = max(abs(v) for v in vals)
ax.set_xlim(-max_abs * 1.4, max_abs * 1.4)
for i, v in enumerate(vals):
    ax.text(v + (max_abs*0.06 if v >= 0 else -max_abs*0.06), i, f'{v:+.1f}%',
            va='center', ha='left' if v >= 0 else 'right', fontsize=9)
plt.tight_layout()
plt.savefig('tornado_interaction_residuals.png', dpi=150)
plt.close()
print("Saved: tornado_interaction_residuals.png")

province_sq2_ratios = []
for prov, d in provinces.items():
    deaths_pre = compute_province_deaths_precode_0417(d)
    deaths_high = compute_province_deaths_highcode_0417(d)
    ratio = deaths_high / deaths_pre
    province_sq2_ratios.append((prov, ratio, d['pga']))
province_sq2_ratios.sort(key=lambda t: t[1])

fig, ax = plt.subplots(figsize=(7.5, 5.5))
labels2 = [f"{t[0]} (PGA {t[2]:.2f})" for t in province_sq2_ratios]
vals2 = [t[1] for t in province_sq2_ratios]
ax.barh(labels2, vals2, color='#eda100')
ax.set_xlabel('High-Code deaths / Pre-Code deaths (lower = bigger benefit from compliance)')
ax.set_title('SQ2 sensitivity by province, ranked (at 04:17), HAZUS')
ax.set_xlim(0, max(vals2) * 1.15)
for i, v in enumerate(vals2):
    ax.text(v + max(vals2)*0.015, i, f'{v:.2f}', va='center', fontsize=8)
plt.tight_layout()
plt.savefig('tornado_province_sq2_sensitivity.png', dpi=150)
plt.close()
print("Saved: tornado_province_sq2_sensitivity.png")

province_sq1_ratios = []
for prov, d in provinces.items():
    ratio = d['occ_day'] / d['occ_night']
    province_sq1_ratios.append((prov, ratio))
province_sq1_ratios.sort(key=lambda t: t[1])

fig, ax = plt.subplots(figsize=(7.5, 5.5))
labels3 = [t[0] for t in province_sq1_ratios]
vals3 = [t[1] for t in province_sq1_ratios]
ax.barh(labels3, vals3, color='#2a78d6')
ax.set_xlabel('Day occupants / Night occupants (lower = bigger day-vs-night swing)')
ax.set_title('SQ1 sensitivity by province, ranked (occupancy-schedule driven)')
ax.set_xlim(0, max(vals3) * 1.15)
for i, v in enumerate(vals3):
    ax.text(v + max(vals3)*0.015, i, f'{v:.3f}', va='center', fontsize=8)
plt.tight_layout()
plt.savefig('tornado_province_sq1_sensitivity.png', dpi=150)
plt.close()
print("Saved: tornado_province_sq1_sensitivity.png")

# ============================================================
# STEP 22: Province-level heatmap across all 9 cells, HAZUS
# ============================================================
from matplotlib.colors import LogNorm

TIME_DESC_ONLY = {'04:17': 'night', '14:00': 'day', '17:00': 'transit'}

cell_order = [
    ('Pre-Code', '04:17'), ('Pre-Code', '14:00'), ('Pre-Code', '17:00'),
    ('Low-Code', '04:17'), ('Low-Code', '14:00'), ('Low-Code', '17:00'),
    ('High-Code', '04:17'), ('High-Code', '14:00'), ('High-Code', '17:00'),
]
cell_fn_map = {
    ('Pre-Code','04:17'): compute_province_deaths_precode_0417,
    ('Pre-Code','14:00'): compute_province_deaths_precode_1400,
    ('Pre-Code','17:00'): compute_province_deaths_precode_1700,
    ('Low-Code','04:17'): compute_province_deaths_lowcode_0417,
    ('Low-Code','14:00'): compute_province_deaths_lowcode_1400,
    ('Low-Code','17:00'): compute_province_deaths_lowcode_1700,
    ('High-Code','04:17'): compute_province_deaths_highcode_0417,
    ('High-Code','14:00'): compute_province_deaths_highcode_1400,
    ('High-Code','17:00'): compute_province_deaths_highcode_1700,
}

province_list = list(provinces.keys())
heat_data = np.zeros((len(province_list), len(cell_order)))
for pi, prov in enumerate(province_list):
    for ci, cell in enumerate(cell_order):
        heat_data[pi, ci] = cell_fn_map[cell](provinces[prov])

sort_idx = np.argsort(-heat_data[:, 0])
heat_data = heat_data[sort_idx, :]
province_list_sorted = [province_list[i] for i in sort_idx]

fig, ax = plt.subplots(figsize=(10, 7))
im = ax.imshow(heat_data, cmap='Blues', aspect='auto', norm=LogNorm(vmin=max(heat_data.min(),0.1), vmax=heat_data.max()))
ax.set_xticks(range(len(cell_order)))
ax.set_xticklabels([f'{e}\n{t}\n({TIME_DESC_ONLY[t]})' for e, t in cell_order], fontsize=8)
ax.set_yticks(range(len(province_list_sorted)))
ax.set_yticklabels(province_list_sorted, fontsize=9)

PRECISION_THRESHOLD = 10
for i in range(heat_data.shape[0]):
    for j in range(heat_data.shape[1]):
        val = heat_data[i, j]
        rel = np.log(val + 1) / np.log(heat_data.max() + 1)
        text_color = 'white' if rel > 0.55 else 'black'
        label = f'{val:,.2f}' if val < PRECISION_THRESHOLD else f'{val:,.0f}'
        ax.text(j, i, label, ha='center', va='center', color=text_color, fontsize=7)

ax.set_title('Deaths by province and scenario cell (log color scale), HAZUS, sorted by Pre-Code x 04:17')
fig.colorbar(im, ax=ax, label='Deaths (log scale)')
plt.tight_layout()
plt.savefig('province_level_heatmap.png', dpi=150)
plt.close()
print("Saved: province_level_heatmap.png")

# ============================================================
# STEP 23: Choropleth maps (PGA and fatalities)
# ============================================================
import geopandas as gpd
import json
import requests

GEOJSON_URL = "https://raw.githubusercontent.com/alpers/Turkey-Maps-GeoJSON/master/tr-cities.json"
resp = requests.get(GEOJSON_URL)
resp.raise_for_status()
geojson_data = resp.json()

gdf_all = gpd.GeoDataFrame.from_features(geojson_data['features'])
gdf_all = gdf_all.set_crs(epsg=4326)

STUDY_PROVINCES = list(provinces.keys())
gdf = gdf_all[gdf_all['name'].isin(STUDY_PROVINCES)].copy()

gdf['deaths_precode_0417'] = gdf['name'].map(lambda p: compute_province_deaths_precode_0417(provinces[p]))

# ============================================================
# STEP 24: Model/Actual ratio map, with epicenters and faults
# ============================================================

ACTUAL_DEATHS = {
    'Adana': 454, 'Adıyaman': 8387, 'Diyarbakır': 414, 'Elazığ': 5,
    'Gaziantep': 3897, 'Hatay': 24147, 'Kahramanmaraş': 12622, 'Kilis': 74,
    'Malatya': 1393, 'Osmaniye': 991, 'Şanlıurfa': 340,
}
gdf['actual'] = gdf['name'].map(ACTUAL_DEATHS)
gdf['ratio'] = gdf['deaths_precode_0417'] / gdf['actual']

epicenters = [
    {'name': 'Pazarcık Mw7.8', 'lon': 37.014, 'lat': 37.226},
    {'name': 'Elbistan Mw7.5', 'lon': 37.203, 'lat': 38.024},
]

FAULTS_URL = "https://raw.githubusercontent.com/GEMScienceTools/gem-global-active-faults/master/geojson/gem_active_faults.geojson"
faults_resp = requests.get(FAULTS_URL, timeout=120)
faults_resp.raise_for_status()
faults_geojson = faults_resp.json()
gdf_faults_all = gpd.GeoDataFrame.from_features(faults_geojson['features']).set_crs(epsg=4326)

minx, miny, maxx, maxy = gdf.total_bounds
buffer = 1.0
gdf_faults = gdf_faults_all.cx[minx-buffer:maxx+buffer, miny-buffer:maxy+buffer]

from shapely.geometry import Point as ShpPoint
mean_lat_eq = sum(e['lat'] for e in epicenters) / len(epicenters)
km_per_deg = 111.32 * np.cos(np.radians(mean_lat_eq))
buffer_deg = 30 / km_per_deg
epicenter_buffers = [ShpPoint(e['lon'], e['lat']).buffer(buffer_deg) for e in epicenters]
near_epicenter_mask = gdf_faults.geometry.apply(lambda geom: any(geom.intersects(b) for b in epicenter_buffers))
gdf_faults = gdf_faults[near_epicenter_mask]

gdf_faults_clipped = gpd.clip(gdf_faults, gdf.unary_union) if len(gdf_faults) > 0 else gdf_faults

# NOTE: the standalone Model/Actual ratio map was replaced by the combined
# HAZUS vs GEM province sensitivity map (STEP 32, below), once GEM's
# per-province death functions become available.

label_offsets = {'Elazığ': (0, -0.25), 'Kahramanmaraş': (-0.25, -0.15), 'Gaziantep': (0.35, 0.08)}

mean_lat = (miny + maxy) / 2
km_per_degree_lon = 111.32 * np.cos(np.radians(mean_lat))
bar_km = 50
bar_deg = bar_km / km_per_degree_lon
bar_x0 = minx + (maxx - minx) * 0.03
bar_y0 = miny - (maxy - miny) * 0.05

# ============================================================
# STEP 25: Simple study area map -- provinces, rupture, epicenters
# (presentation figure, no data overlay)
# ============================================================

fig, ax = plt.subplots(figsize=(9, 7.5))
gdf.plot(ax=ax, color='#1a3a6b', edgecolor='#B4B2A9', linewidth=1.0)

for idx, row in gdf.iterrows():
    centroid = row.geometry.centroid
    dx, dy = label_offsets.get(row['name'], (0, 0))
    ax.annotate(row['name'], (centroid.x + dx, centroid.y + dy), ha='center', fontsize=8, color='white')

if len(gdf_faults_clipped) > 0:
    gdf_faults_clipped.plot(ax=ax, color='#800000', linewidth=2.0, alpha=0.9, zorder=3, label='Fault rupture')

for eq in epicenters:
    ax.plot(eq['lon'], eq['lat'], marker='*', color='yellow', markeredgecolor='black', markersize=20, zorder=5)
    ax.annotate(eq['name'], (eq['lon'], eq['lat']), textcoords="offset points", xytext=(8, 8),
                fontsize=9, zorder=6, color='white', fontweight='medium')

ax.plot([bar_x0, bar_x0 + bar_deg], [bar_y0, bar_y0], color='black', linewidth=2, zorder=6)
ax.plot([bar_x0, bar_x0], [bar_y0 - 0.05, bar_y0 + 0.05], color='black', linewidth=1.5, zorder=6)
ax.plot([bar_x0 + bar_deg, bar_x0 + bar_deg], [bar_y0 - 0.05, bar_y0 + 0.05], color='black', linewidth=1.5, zorder=6)
ax.annotate(f'{bar_km} km', (bar_x0 + bar_deg/2, bar_y0 - 0.15), ha='center', fontsize=8, zorder=6)

from matplotlib.lines import Line2D
legend_elements = [
    Line2D([0], [0], color='#800000', linewidth=2.0, label='Fault rupture'),
    Line2D([0], [0], marker='*', color='w', markerfacecolor='yellow', markeredgecolor='black',
           markersize=14, label='Earthquake epicenter'),
]
ax.legend(handles=legend_elements, loc='lower left', fontsize=9, frameon=True)

ax.set_title('Study area: 11 earthquake-affected provinces\nwith fault rupture and epicenters (2023 Türkiye-Syria earthquakes)', fontsize=11, pad=10)
ax.set_axis_off()
ax.set_xlim(minx - (maxx-minx)*0.1, maxx + (maxx-minx)*0.1)
ax.set_ylim(miny - (maxy-miny)*0.15, maxy + (maxy-miny)*0.15)

plt.tight_layout()
plt.savefig('map_study_area.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: map_study_area.png")

# ============================================================
# STEP 26: SQ2 sensitivity data (HAZUS)
# (standalone map removed; reused by the HAZUS vs GEM map, STEP 32)
# ============================================================

gdf['deaths_highcode_0417'] = gdf['name'].map(lambda p: compute_province_deaths_highcode_0417(provinces[p]))
gdf['pct_reduction'] = (1 - gdf['deaths_highcode_0417'] / gdf['deaths_precode_0417']) * 100

# ============================================================
# STEP 27: GEM recomputation with corrected PGA
# ============================================================

gem_lookup = {}
for row in wb['Fatalities Data (Long)'].iter_rows(min_row=2, values_only=True):
    func_id, taxonomy, occupancy, dist, pt, pga_val, meanLR, covLR = row
    parts = taxonomy.split('/')
    material = parts[0]
    combo = next((p for p in parts if 'CD' in p), None)
    cd = combo.split('+')[0] if combo else None
    height = int(taxonomy.split('H:')[1].split('/')[0])
    key = (material, cd, height, occupancy)
    gem_lookup.setdefault(key, {'pga': [], 'lr': []})
    gem_lookup[key]['pga'].append(pga_val)
    gem_lookup[key]['lr'].append(meanLR)

def gem_death_fraction(material, cd, height, pga, occ='RES'):
    key = (material, cd, height, occ)
    if key not in gem_lookup:
        return None
    p = np.array(gem_lookup[key]['pga']); l = np.array(gem_lookup[key]['lr'])
    o = np.argsort(p)
    return float(np.interp(pga, p[o], l[o]))

MASONRY_SUBS = ['MUR+ADO', 'MUR+CBH', 'MUR+CLBRS', 'MUR+STDRE', 'MUR+STRUB']
def gem_masonry_fraction(cd, height, pga):
    vals = [gem_death_fraction(s, cd, height, pga) for s in MASONRY_SUBS]
    vals = [v for v in vals if v is not None]
    return float(np.mean(vals)) if vals else None

GEM_MAP = {
    'Pre-Code':  {'CR': 'CDL', 'MAS': 'CDN', 'W': 'CDN', 'S': 'CDL'},
    'Low-Code':  {'CR': 'CDM', 'MAS': 'CDN', 'W': 'CDM', 'S': 'CDM'},
    'High-Code': {'CR': 'CDH', 'MAS': 'CDN', 'W': 'CDM', 'S': 'CDH'},
}

GEM_HEIGHT_OPTIONS = np.array([1, 2, 3, 4, 5, 6, 8, 10])

def snap_to_gem_height(x):
    return int(GEM_HEIGHT_OPTIONS[np.argmin(np.abs(GEM_HEIGHT_OPTIONS - x))])

ws_floor = wb['RAW Floor Distribution']

floor_header_row_idx = None
for row_cells in ws_floor.iter_rows(min_row=1, max_row=15):
    if row_cells[0].value == 'Province':
        floor_header_row_idx = row_cells[0].row
        break

if floor_header_row_idx is None:
    raise ValueError("Could not find the 'Province' header row in 'RAW Floor Distribution'.")

for row in ws_floor.iter_rows(min_row=floor_header_row_idx + 1, max_row=floor_header_row_idx + 30, values_only=True):
    prov = row[0]
    if prov not in provinces:
        continue
    f1, f2, f3, f4, f5, f6plus = row[3], row[4], row[5], row[6], row[7], row[8]
    low_n = f1 + f2 + f3
    low_avg = (1*f1 + 2*f2 + 3*f3) / low_n if low_n else 2
    mid_n = f4 + f5
    mid_avg = (4*f4 + 5*f5) / mid_n if mid_n else 4
    high_avg = 6
    provinces[prov]['gem_height'] = {
        'low': snap_to_gem_height(low_avg),
        'mid': snap_to_gem_height(mid_avg),
        'high': snap_to_gem_height(high_avg),
    }

def compute_gem_deaths(prov_data, era, occ_key):
    pga = prov_data['pga']
    occ = prov_data[occ_key]
    total = 0.0
    heights = {'low': prov_data['height_low'], 'mid': prov_data['height_mid'], 'high': prov_data['height_high']}
    m = GEM_MAP[era]
    gem_h = prov_data['gem_height']

    for hband, hfrac in heights.items():
        h = gem_h[hband]
        lr = gem_death_fraction('CR', m['CR'], h, pga)
        if lr is not None:
            total += occ * prov_data['rc_bldg_frac'] * hfrac * lr
        lr = gem_masonry_fraction(m['MAS'], h, pga)
        if lr is not None:
            total += occ * prov_data['mas_bldg_frac'] * hfrac * lr
        lr = gem_death_fraction('S', m['S'], h, pga)
        if lr is not None:
            total += occ * prov_data['steel_bldg_frac'] * hfrac * lr

    lr = gem_death_fraction('W', m['W'], gem_h['low'], pga)
    if lr is not None:
        total += occ * prov_data['timber_bldg_frac'] * lr

    return total

# ============================================================
# STEP 28: Full 9-cell GEM matrix
# ============================================================

SQ2_ERAS = ['Pre-Code', 'Low-Code', 'High-Code']
SQ1_OCC_KEYS = {'04:17': 'occ_night', '14:00': 'occ_day', '17:00': 'occ_transit'}

gem_matrix = {}
for era in SQ2_ERAS:
    for time_label, occ_key in SQ1_OCC_KEYS.items():
        total = sum(compute_gem_deaths(d, era, occ_key) for d in provinces.values())
        gem_matrix[(era, time_label)] = total

print("\nGEM: National Total Deaths per Cell")
print(f"{'':12} {'04:17':>12} {'14:00':>12} {'17:00':>12}")
for era in SQ2_ERAS:
    row = [gem_matrix[(era, t)] for t in ['04:17', '14:00', '17:00']]
    print(f"{era:12} {row[0]:12,.1f} {row[1]:12,.1f} {row[2]:12,.1f}")

hazus_matrix = {
    ('Pre-Code','04:17'): national_total, ('Pre-Code','14:00'): national_total_1400, ('Pre-Code','17:00'): national_total_1700,
    ('Low-Code','04:17'): national_total_lowcode_0417, ('Low-Code','14:00'): national_total_lowcode_1400, ('Low-Code','17:00'): national_total_lowcode_1700,
    ('High-Code','04:17'): national_total_highcode_0417, ('High-Code','14:00'): national_total_highcode_1400, ('High-Code','17:00'): national_total_highcode_1700,
}

print("\nGEM vs HAZUS vs Actual (all 9 cells)")
print(f"{'Cell':22} {'GEM':>12} {'HAZUS':>12} {'HAZUS/GEM':>12}")
for era in SQ2_ERAS:
    for t in ['04:17', '14:00', '17:00']:
        g = gem_matrix[(era, t)]
        h = hazus_matrix[(era, t)]
        ratio = h / g if g > 0 else float('inf')
        print(f"{era+' x '+t:22} {g:12,.1f} {h:12,.1f} {ratio:12,.0f}")

def compute_province_deaths_factual_gem_0417(prov_data):
    deaths_pre  = compute_gem_deaths(prov_data, 'Pre-Code', 'occ_night')
    deaths_low  = compute_gem_deaths(prov_data, 'Low-Code', 'occ_night')
    deaths_high = compute_gem_deaths(prov_data, 'High-Code', 'occ_night')
    return (
        prov_data['precode_pct']  * deaths_pre +
        prov_data['lowcode_pct']  * deaths_low +
        prov_data['highcode_pct'] * deaths_high
    )

national_total_factual_gem = sum(compute_province_deaths_factual_gem_0417(d) for d in provinces.values())

print(f"\nNational total (GEM FACTUAL, actual code-era mix x 04:17): {national_total_factual_gem:,.2f}")
print(f"National total (HAZUS FACTUAL):                             {national_total_factual:,.1f}")
print(f"GEM Factual / Actual ratio:   {national_total_factual_gem/52724:.4f}")
print(f"HAZUS Factual / Actual ratio: {national_total_factual/52724:.3f}")
print(f"HAZUS Factual / GEM Factual ratio: {national_total_factual/national_total_factual_gem:,.0f}x")

# ============================================================
# STEP 29: National total comparison chart -- HAZUS vs GEM vs Chen
# ============================================================

fig, ax = plt.subplots(figsize=(8, 5.5))
labels_nat = ['HAZUS\nFactual', 'GEM\nFactual', 'Chen et al. (2025)\nactual reported']
values_nat = [national_total_factual, national_total_factual_gem, 52724]
colors_nat = ['#2a78d6', '#3aa655', '#eda100']

bars_nat = ax.bar(labels_nat, values_nat, color=colors_nat)
ax.set_yscale('log')
ax.set_ylabel('National total deaths (log scale)')
ax.set_ylim(top=max(values_nat) * 3)
ax.set_axisbelow(True)
ax.grid(axis='y', which='major', alpha=0.3)
for bar, val in zip(bars_nat, values_nat):
    ax.annotate(f'{val:,.0f}', xy=(bar.get_x() + bar.get_width()/2, val),
                xytext=(0, 5), textcoords='offset points', ha='center', fontsize=10)
plt.tight_layout()
plt.savefig('national_total_comparison.png', dpi=150)
plt.close()
print("Saved: national_total_comparison.png")

# ============================================================
# STEP 30: Compliance comparison at 04:17 -- uniform bounds vs
# realistic (factual) baseline, HAZUS vs GEM
# ============================================================

from matplotlib.ticker import LogLocator

compliance_labels = ['Uniform\nPre-Code', 'Factual\n(2023 mix)', 'Uniform\nHigh-Code']
compliance_hazus = [national_total, national_total_factual, national_total_highcode_0417]
compliance_gem = [gem_matrix[('Pre-Code', '04:17')], national_total_factual_gem, gem_matrix[('High-Code', '04:17')]]

x_comp = np.arange(3)
width_comp = 0.35

fig, ax = plt.subplots(figsize=(8, 5.5))
b1 = ax.bar(x_comp - width_comp/2, compliance_hazus, width_comp, label='HAZUS', color='#2a78d6')
b2 = ax.bar(x_comp + width_comp/2, compliance_gem, width_comp, label='GEM', color='#3aa655')

ax.set_yscale('log')
ax.set_ylabel('National total deaths (log scale)')
ax.set_title('Compliance comparison at 04:17: uniform bounds vs realistic (factual) baseline')
ax.set_xticks(x_comp)
ax.set_xticklabels(compliance_labels)
ax.legend()
ax.set_axisbelow(True)
ax.yaxis.set_major_locator(LogLocator(base=10.0, numticks=4))
ax.grid(axis='y', which='major', alpha=0.3)

def fmt_comp(h):
    return f'{h:.3f}' if h < 1 else f'{h:,.0f}'

for rects in [b1, b2]:
    for rect in rects:
        h = rect.get_height()
        ax.annotate(fmt_comp(h), xy=(rect.get_x() + rect.get_width()/2, h),
                    xytext=(0, 5), textcoords='offset points', ha='center', fontsize=9)

ax.set_ylim(top=max(compliance_hazus) * 15)

uniform_ratio_hazus = compliance_hazus[0] / compliance_hazus[2]
uniform_ratio_gem = compliance_gem[0] / compliance_gem[2]
realistic_ratio_hazus = compliance_hazus[1] / compliance_hazus[2]
realistic_ratio_gem = compliance_gem[1] / compliance_gem[2]

y_uniform_arrow = max(compliance_hazus) * 6.5
y_realistic_arrow = max(compliance_hazus) * 2.2

ax.annotate('', xy=(x_comp[0] - width_comp/2, y_uniform_arrow), xytext=(x_comp[2] - width_comp/2, y_uniform_arrow),
            arrowprops=dict(arrowstyle='<->', color='#8b1a1a', lw=1.5))
ax.text((x_comp[0] + x_comp[2]) / 2 - width_comp/2, y_uniform_arrow * 1.15,
        f'Uniform ratio: {uniform_ratio_hazus:.2f}x (HAZUS) / {uniform_ratio_gem:.2f}x (GEM)',
        ha='center', color='#8b1a1a', fontsize=9)

ax.annotate('', xy=(x_comp[1] - width_comp/2, y_realistic_arrow), xytext=(x_comp[2] - width_comp/2, y_realistic_arrow),
            arrowprops=dict(arrowstyle='<->', color='#2a78d6', lw=1.5))
ax.text((x_comp[1] + x_comp[2]) / 2 - width_comp/2, y_realistic_arrow * 1.2,
        f'Realistic ratio: {realistic_ratio_hazus:.3f}x (HAZUS) / {realistic_ratio_gem:.3f}x (GEM)',
        ha='center', color='#2a78d6', fontsize=9)

ax.annotate('Uniform ratio = Uniform Pre-Code / Uniform High-Code deaths, per model.\n'
            'Realistic ratio = Factual (2023 compliance mix) / Uniform High-Code deaths, per model.\n'
            'All bars computed at 04:17.',
            xy=(0.5, -0.32), xycoords='axes fraction', ha='center', fontsize=8, color='#444444')

plt.subplots_adjust(bottom=0.36)
plt.savefig('compliance_comparison_uniform_vs_realistic.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: compliance_comparison_uniform_vs_realistic.png")

# ============================================================
# STEP 31: Share of theoretical compliance benefit already
# captured by the 2023 mix, HAZUS vs GEM
# ============================================================

theoretical_benefit_hazus = compliance_hazus[0] - compliance_hazus[2]
captured_benefit_hazus = compliance_hazus[0] - compliance_hazus[1]
pct_captured_hazus = captured_benefit_hazus / theoretical_benefit_hazus * 100
pct_remaining_hazus = 100 - pct_captured_hazus

theoretical_benefit_gem = compliance_gem[0] - compliance_gem[2]
captured_benefit_gem = compliance_gem[0] - compliance_gem[1]
pct_captured_gem = captured_benefit_gem / theoretical_benefit_gem * 100
pct_remaining_gem = 100 - pct_captured_gem

models_benefit = ['HAZUS', 'GEM']
pct_captured_vals = [pct_captured_hazus, pct_captured_gem]
pct_remaining_vals = [pct_remaining_hazus, pct_remaining_gem]

fig, ax = plt.subplots(figsize=(6.5, 5.5))
bars_captured = ax.bar(models_benefit, pct_captured_vals, color='#2a78d6', label='Already captured (2023 mix)')
bars_remaining = ax.bar(models_benefit, pct_remaining_vals, bottom=pct_captured_vals, color='#c9d9ef',
                         label='Remaining (achievable via retrofit)')

for i, (cap, rem) in enumerate(zip(pct_captured_vals, pct_remaining_vals)):
    ax.text(i, cap / 2, f'{cap:.1f}%', ha='center', va='center', fontsize=11, color='white', fontweight='bold')
    ax.text(i, cap + rem / 2, f'{rem:.1f}%', ha='center', va='center', fontsize=11, color='#333333')

ax.set_ylabel('% of theoretical compliance benefit')
ax.set_title('Share of theoretical compliance benefit\nalready captured by the 2023 compliance mix')
ax.set_ylim(0, 118)
ax.set_yticks([0, 20, 40, 60, 80, 100])
ax.legend(loc='upper center', ncol=2, frameon=False, fontsize=8)

ax.text(0.5, -0.2,
        'Theoretical benefit = Uniform Pre-Code − Uniform High-Code deaths, per model.\n'
        'Already captured = Uniform Pre-Code − Factual (2023 mix) deaths, per model.\n'
        'All values computed at 04:17.',
        transform=ax.transAxes, ha='center', fontsize=8, color='#444444')

plt.subplots_adjust(top=0.85, bottom=0.28)
plt.savefig('compliance_benefit_captured.png', dpi=150)
plt.close()
print("Saved: compliance_benefit_captured.png")

# ============================================================
# STEP 32: Province SQ2 sensitivity map -- HAZUS vs GEM, side by side
# (replaces the old standalone Model/Actual ratio map)
# ============================================================

gdf['deaths_precode_0417_gem'] = gdf['name'].map(lambda p: compute_gem_deaths(provinces[p], 'Pre-Code', 'occ_night'))
gdf['deaths_highcode_0417_gem'] = gdf['name'].map(lambda p: compute_gem_deaths(provinces[p], 'High-Code', 'occ_night'))
gdf['pct_reduction_gem'] = (1 - gdf['deaths_highcode_0417_gem'] / gdf['deaths_precode_0417_gem']) * 100

def draw_sensitivity_panel(ax, column, title, cmap, eq_label_color='#800000', eq_label_offsets=None,
                            legend_label='% reduction in deaths (Pre-Code -> High-Code)'):
    gdf.plot(column=column, cmap=cmap, linewidth=0.6, edgecolor='black', legend=True,
             ax=ax, legend_kwds={'label': legend_label, 'shrink': 0.7},
             vmin=0, vmax=100)
    if len(gdf_faults_clipped) > 0:
        gdf_faults_clipped.plot(ax=ax, color='#800000', linewidth=1.2, alpha=0.9, zorder=3)
    for eq in epicenters:
        ax.plot(eq['lon'], eq['lat'], marker='*', color='yellow', markeredgecolor='black', markersize=14, zorder=5)
        dx, dy = (eq_label_offsets or {}).get(eq['name'], (8, 8))
        ha = 'right' if dx < 0 else 'left'
        ax.annotate(eq['name'], (eq['lon'], eq['lat']), textcoords="offset points", xytext=(dx, dy),
                    ha=ha, fontsize=7, zorder=6, color=eq_label_color)

    for idx, row in gdf.iterrows():
        centroid = row.geometry.centroid
        label = f"{row['name']}\n{row[column]:.0f}%"
        rel_val = row[column] / 100
        text_color = 'white' if rel_val > 0.55 else 'black'
        dx, dy = label_offsets.get(row['name'], (0, 0))
        ax.annotate(label, (centroid.x + dx, centroid.y + dy), ha='center', fontsize=6.5, color=text_color, zorder=4)

    ax.plot([bar_x0, bar_x0 + bar_deg], [bar_y0, bar_y0], color='black', linewidth=2, zorder=6)
    ax.plot([bar_x0, bar_x0], [bar_y0 - 0.05, bar_y0 + 0.05], color='black', linewidth=1.5, zorder=6)
    ax.plot([bar_x0 + bar_deg, bar_x0 + bar_deg], [bar_y0 - 0.05, bar_y0 + 0.05], color='black', linewidth=1.5, zorder=6)
    ax.annotate(f'{bar_km} km', (bar_x0 + bar_deg/2, bar_y0 - 0.15), ha='center', fontsize=7, zorder=6)

    ax.set_title(title, fontsize=10, pad=10)
    ax.set_axis_off()
    ax.set_ylim(miny - (maxy-miny)*0.15, maxy + (maxy-miny)*0.35)

fig, axes = plt.subplots(1, 2, figsize=(17, 8))
draw_sensitivity_panel(axes[0], 'pct_reduction', 'HAZUS: % reduction, Pre-Code -> High-Code (04:17)', 'Blues',
                        eq_label_color='yellow',
                        eq_label_offsets={'Pazarcık Mw7.8': (8, -16), 'Elbistan Mw7.5': (-3, 8)})
draw_sensitivity_panel(axes[1], 'pct_reduction_gem', 'GEM: % reduction, Pre-Code -> High-Code (04:17)', 'Greens',
                        eq_label_offsets={'Pazarcık Mw7.8': (8, -16), 'Elbistan Mw7.5': (-3, 8)})
fig.suptitle('Province-level compliance sensitivity: HAZUS vs GEM', fontsize=13, y=0.98)
plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.savefig('map_sq2_sensitivity_hazus_vs_gem.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: map_sq2_sensitivity_hazus_vs_gem.png")

# ============================================================
# STEP 33: Factual scenario deaths by province chart -- HAZUS vs GEM vs Chen
# ============================================================

factual_chart_data = []
for prov, d in provinces.items():
    hf = compute_province_deaths_factual_0417(d)
    gf = compute_province_deaths_factual_gem_0417(d)
    chen = ACTUAL_DEATHS[prov]
    factual_chart_data.append((prov, hf, gf, chen))

factual_chart_data.sort(key=lambda t: t[1], reverse=True)
chart_provinces = [t[0] for t in factual_chart_data]
chart_hazus = [t[1] for t in factual_chart_data]
chart_gem = [t[2] for t in factual_chart_data]
chart_chen = [t[3] for t in factual_chart_data]

x = np.arange(len(chart_provinces))
width = 0.27

fig, ax = plt.subplots(figsize=(12, 6.5))
b1 = ax.bar(x - width, chart_hazus, width, label='HAZUS Factual', color='#1f77b4')
b2 = ax.bar(x,         chart_gem,   width, label='GEM Factual',   color='#2ca02c')
b3 = ax.bar(x + width,  chart_chen, width, label='Chen et al. (2025)', color='#eda100')

ax.set_yscale('log')
ax.set_ylim(top=max(chart_hazus + chart_chen) * 10)
ax.set_ylabel('Deaths (log scale)')
ax.set_title('Factual scenario deaths by province (04:17): HAZUS vs GEM vs Chen et al. (2025)')
ax.set_xticks(x)
ax.set_xticklabels(chart_provinces, rotation=35, ha='right')
ax.legend()
ax.set_axisbelow(True)
ax.grid(axis='y', which='major', alpha=0.3)

def fmt_bar(h):
    if h < 1:
        return f'{h:.3f}'
    elif h < 100:
        return f'{h:,.1f}'
    else:
        return f'{h:,.0f}'

for rects in [b1, b2, b3]:
    for rect in rects:
        h = rect.get_height()
        if h > 0:
            ax.annotate(fmt_bar(h), xy=(rect.get_x() + rect.get_width() / 2, h),
                        xytext=(0, 3), textcoords="offset points",
                        ha='center', va='bottom', fontsize=7, rotation=90)

plt.tight_layout()
plt.savefig('factual_province_breakdown.png', dpi=200)
plt.close()
print("Saved: factual_province_breakdown.png (HAZUS vs GEM vs Chen)")

# ============================================================
# STEP 34: Retrofit sweep -- GEM version, all 3 timing scenarios
# ============================================================

def compute_province_deaths_retrofit_gem(prov_data, r, time_label):
    occ_key = SQ1_OCC_KEYS[time_label]
    deaths_pre  = compute_gem_deaths(prov_data, 'Pre-Code', occ_key)
    deaths_low  = compute_gem_deaths(prov_data, 'Low-Code', occ_key)
    deaths_high = compute_gem_deaths(prov_data, 'High-Code', occ_key)

    pre_pct  = prov_data['precode_pct']  * (1 - r)
    low_pct  = prov_data['lowcode_pct']  * (1 - r)
    high_pct = prov_data['highcode_pct'] + r * (prov_data['precode_pct'] + prov_data['lowcode_pct'])

    return pre_pct * deaths_pre + low_pct * deaths_low + high_pct * deaths_high

retrofit_national_gem_by_time = {t: {} for t in ['04:17', '14:00', '17:00']}
for time_label in ['04:17', '14:00', '17:00']:
    for r in RETROFIT_LEVELS:
        total = sum(compute_province_deaths_retrofit_gem(d, r, time_label) for d in provinces.values())
        retrofit_national_gem_by_time[time_label][r] = total

df_retrofit_gem_alltime = pd.DataFrame(
    {t: [retrofit_national_gem_by_time[t][r] for r in RETROFIT_LEVELS] for t in ['04:17', '14:00', '17:00']},
    index=[f"{r*100:.0f}%" for r in RETROFIT_LEVELS]
)
df_retrofit_gem_alltime.index.name = 'Retrofit %'
df_retrofit_gem_alltime.to_excel('retrofit_sweep_gem_all_times.xlsx')
print("\nSaved: retrofit_sweep_gem_all_times.xlsx")

print("\n" + "="*70)
print("Diminishing-returns check across all 3 timings, GEM (Figure 12 data)")
print("="*70)
for time_label in ['04:17', '14:00', '17:00']:
    print_retrofit_targets(retrofit_national_gem_by_time, time_label, 'GEM')

print("\n" + "="*70)
print("HAZUS vs GEM: maximum achievable retrofit ceiling, all 3 timings (Figure 12)")
print("="*70)
for time_label in ['04:17', '14:00', '17:00']:
    hazus_baseline = retrofit_national_by_time[time_label][0.0]
    hazus_floor = min(retrofit_national_by_time[time_label].values())
    hazus_ceiling_pct = (1 - hazus_floor / hazus_baseline) * 100

    gem_baseline = retrofit_national_gem_by_time[time_label][0.0]
    gem_floor = min(retrofit_national_gem_by_time[time_label].values())
    gem_ceiling_pct = (1 - gem_floor / gem_baseline) * 100

    print(f"  {time_label:8} HAZUS max reduction: {hazus_ceiling_pct:5.1f}%   "
          f"GEM max reduction: {gem_ceiling_pct:5.1f}%")

fig, axes = plt.subplots(1, 3, figsize=(15, 5.5), sharey=True)
for ax, time_label in zip(axes, ['04:17', '14:00', '17:00']):
    hazus_vals = [(1 - retrofit_national_by_time[time_label][r] / retrofit_national_by_time[time_label][0.0]) * 100
                  for r in RETROFIT_LEVELS]
    gem_vals = [(1 - retrofit_national_gem_by_time[time_label][r] / retrofit_national_gem_by_time[time_label][0.0]) * 100
                for r in RETROFIT_LEVELS]
    ax.plot(r_pct, hazus_vals, marker='o', markersize=4, linewidth=2.5, color='#2a78d6', label='HAZUS')
    ax.plot(r_pct, gem_vals, marker='o', markersize=4, linewidth=2.5, color='#3aa655', label='GEM')

    for lp in [0, 20, 40, 60, 80, 100]:
        r_key = round(lp / 100, 2)
        hazus_at_lp = (1 - retrofit_national_by_time[time_label][r_key] / retrofit_national_by_time[time_label][0.0]) * 100
        gem_at_lp = (1 - retrofit_national_gem_by_time[time_label][r_key] / retrofit_national_gem_by_time[time_label][0.0]) * 100
        ax.annotate(f'{hazus_at_lp:.1f}%', xy=(lp, hazus_at_lp), xytext=(0, 7),
                    textcoords='offset points', ha='center', fontsize=7, color='#2a78d6', fontweight='medium')
        ax.annotate(f'{gem_at_lp:.1f}%', xy=(lp, gem_at_lp), xytext=(0, -12),
                    textcoords='offset points', ha='center', fontsize=7, color='#3aa655', fontweight='medium')

    ax.set_xlabel('Retrofit coverage')
    ax.set_title(TIME_LABELS_DISPLAY[time_label])
    ax.set_xticks([0, 20, 40, 60, 80, 100])
    ax.set_xticklabels(['0%', '20%', '40%', '60%', '80%', '100%'])
    ax.set_axisbelow(True)
    ax.grid(alpha=0.3)

axes[0].set_ylabel('Fatality reduction (%)')
axes[0].legend(loc='upper left', frameon=False, fontsize=9)
fig.suptitle('Fatality reduction vs. % non-compliant stock retrofitted, by event timing: HAZUS vs GEM', fontsize=13)
plt.tight_layout()
plt.savefig('fatality_reduction_vs_retrofit_hazus_gem.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: fatality_reduction_vs_retrofit_hazus_gem.png")

fig, ax = plt.subplots(figsize=(8, 5.5))
for time_label in ['04:17', '14:00', '17:00']:
    vals = [retrofit_national_gem_by_time[time_label][r] for r in RETROFIT_LEVELS]
    ax.plot(r_pct, vals, marker='o', markersize=3, linewidth=2,
            color=colors_time[time_label], label=time_label)
ax.set_xlabel('Retrofit share of non-High-Code stock upgraded to High-Code (%)')
ax.set_ylabel('National total deaths (GEM)')
ax.set_title('GEM retrofit sensitivity, all 3 timing scenarios')
ax.legend(title='Event time', fontsize=9)
ax.grid(alpha=0.3)
plt.tight_layout()
plt.savefig('retrofit_sweep_gem_all_times.png', dpi=150)
plt.close()
print("Saved: retrofit_sweep_gem_all_times.png")

# ============================================================
# STEP 35: Factual scenario deaths by event timing -- HAZUS vs GEM
# (2023 compliance mix held fixed)
# ============================================================

factual_by_time_hazus = [national_total_factual, national_total_factual_1400, national_total_factual_1700]
factual_by_time_gem = [retrofit_national_gem_by_time[t][0.0] for t in ['04:17', '14:00', '17:00']]

x_evt = np.arange(3)
width_evt = 0.35

fig, ax = plt.subplots(figsize=(8, 5.5))
b1 = ax.bar(x_evt - width_evt/2, factual_by_time_hazus, width_evt, label='HAZUS Factual', color='#2a78d6')
b2 = ax.bar(x_evt + width_evt/2, factual_by_time_gem, width_evt, label='GEM Factual', color='#3aa655')

ax.set_yscale('log')
ax.set_ylabel('National total deaths (log scale)')
ax.set_title('Factual scenario deaths by event timing, HAZUS vs GEM\n(2023 compliance mix held fixed)')
ax.set_xticks(x_evt)
ax.set_xticklabels([TIME_LABELS_DISPLAY[t] for t in ['04:17', '14:00', '17:00']])
ax.legend()
from matplotlib.ticker import LogLocator
ax.set_axisbelow(True)
ax.yaxis.set_major_locator(LogLocator(base=10.0, numticks=4))
ax.grid(axis='y', which='major', alpha=0.3)

def fmt_evt(h):
    if h < 1:
        return f'{h:.3f}'
    elif h < 100:
        return f'{h:,.2f}'
    else:
        return f'{h:,.1f}'

for rects in [b1, b2]:
    for rect in rects:
        h = rect.get_height()
        ax.annotate(fmt_evt(h), xy=(rect.get_x() + rect.get_width()/2, h),
                    xytext=(0, 5), textcoords='offset points', ha='center', fontsize=9)

ax.set_ylim(top=max(factual_by_time_hazus) * 15)

hazus_worst_best_ratio = factual_by_time_hazus[0] / factual_by_time_hazus[1]
gem_worst_best_ratio = factual_by_time_gem[0] / factual_by_time_gem[1]

y_hazus_arrow = max(factual_by_time_hazus) * 6.5
y_gem_arrow = max(factual_by_time_hazus) * 2.2

ax.annotate('', xy=(x_evt[0] - width_evt/2, y_hazus_arrow), xytext=(x_evt[1] - width_evt/2, y_hazus_arrow),
            arrowprops=dict(arrowstyle='<->', color='#2a78d6', lw=1.5))
ax.text((x_evt[0] + x_evt[1]) / 2 - width_evt/2, y_hazus_arrow * 1.15,
        f'HAZUS: {hazus_worst_best_ratio:.2f}x', ha='center', color='#2a78d6', fontsize=9)

ax.annotate('', xy=(x_evt[0] + width_evt/2, y_gem_arrow), xytext=(x_evt[1] + width_evt/2, y_gem_arrow),
            arrowprops=dict(arrowstyle='<->', color='#3aa655', lw=1.5))
ax.text((x_evt[0] + x_evt[1]) / 2 + width_evt/2, y_gem_arrow * 1.2,
        f'GEM: {gem_worst_best_ratio:.2f}x', ha='center', color='#3aa655', fontsize=9)

ax.annotate('Worst point in\nboth models', xy=(x_evt[0] - width_evt/2, factual_by_time_hazus[0] * 1.6),
            xytext=(x_evt[0] - width_evt/2, y_hazus_arrow * 0.5),
            ha='center', fontsize=8, color='#8b1a1a',
            arrowprops=dict(arrowstyle='->', color='#8b1a1a', lw=1.2))

ax.annotate('Ratio = worst-timing deaths (04:17) / best-timing deaths (14:00), per model—'
            'both computed with the 2023 compliance mix held fixed',
            xy=(0.5, -0.16), xycoords='axes fraction', ha='center', fontsize=8, color='#444444')

plt.subplots_adjust(bottom=0.2)
plt.savefig('factual_by_event_timing_hazus_gem.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: factual_by_event_timing_hazus_gem.png")

# ============================================================
# STEP 36: Combined SQ1 + SQ2 sensitivity chart -- GEM
# ============================================================

fig, ax = plt.subplots(figsize=(9.5, 6.5))
for time_label in ['04:17', '14:00', '17:00']:
    vals_pct = [(1 - retrofit_national_gem_by_time[time_label][r] / national_total_factual_gem) * 100 for r in RETROFIT_LEVELS]
    ax.plot(r_pct, vals_pct, marker='o', markersize=3, linewidth=2,
            color=colors_time[time_label], label=time_label)
    for lp in LABEL_POINTS:
        r_key = round(lp / 100, 2)
        val_at_lp = (1 - retrofit_national_gem_by_time[time_label][r_key] / national_total_factual_gem) * 100
        ax.annotate(f'{val_at_lp:.1f}%', xy=(lp, val_at_lp), xytext=(0, 7),
                    textcoords='offset points', ha='center', fontsize=8,
                    color=colors_time[time_label], fontweight='medium')

ax.axhline(0, color='gray', linestyle=':', linewidth=0.8)
ax.annotate('Factual (04:17, r=0) = 0% reduction', xy=(2, 2), fontsize=8, color='gray')
ax.set_xticks(LABEL_POINTS)
ax.set_xlabel('Retrofit share of non-High-Code stock upgraded to High-Code (%)')
ax.set_ylabel('% reduction in deaths, relative to GEM Factual (04:17, r=0) baseline')
ax.set_title('GEM: combined SQ1 (timing) + SQ2 (retrofit) sensitivity\n'
              '(vertical gap at r=0 = SQ1 effect; slope of each line = SQ2 effect)')
ax.legend(title='Event time', fontsize=9)
ax.grid(alpha=0.3)
plt.tight_layout()
plt.savefig('sensitivity_combined_gem.png', dpi=150)
plt.close()
print("Saved: sensitivity_combined_gem.png")

# ============================================================
# STEP 37: Retrofit sweep -- GEM, per-province breakdown (04:17)
# ============================================================

retrofit_gem_by_province = {r: {} for r in RETROFIT_LEVELS}
for r in RETROFIT_LEVELS:
    for prov, d in provinces.items():
        retrofit_gem_by_province[r][prov] = compute_province_deaths_retrofit_gem(d, r, '04:17')

df_retrofit_gem_province = pd.DataFrame(
    {prov: [retrofit_gem_by_province[r][prov] for r in RETROFIT_LEVELS] for prov in provinces},
    index=[f"{r*100:.0f}%" for r in RETROFIT_LEVELS]
)
df_retrofit_gem_province.index.name = 'Retrofit %'
df_retrofit_gem_province['NATIONAL TOTAL'] = [retrofit_national_gem_by_time['04:17'][r] for r in RETROFIT_LEVELS]
df_retrofit_gem_province.to_excel('retrofit_sweep_gem_by_province.xlsx')
print("\nSaved: retrofit_sweep_gem_by_province.xlsx")

# ============================================================
# STEP 38: Retrofit ceiling map -- HAZUS vs GEM, side by side
# (% reduction achievable from Factual 2023 mix -> full retrofit
# to High-Code, at 04:17 -- the policy-relevant number, not the
# theoretical Uniform Pre-Code -> Uniform High-Code range shown
# in the STEP 32 map)
# ============================================================

def retrofit_ceiling(retrofit_dict, prov):
    baseline = retrofit_dict[0.0][prov]
    floor = min(retrofit_dict[r][prov] for r in RETROFIT_LEVELS)
    return (1 - floor / baseline) * 100 if baseline > 0 else float('nan')

gdf['retrofit_ceiling_hazus'] = gdf['name'].map(lambda p: retrofit_ceiling(retrofit_by_province, p))
gdf['retrofit_ceiling_gem'] = gdf['name'].map(lambda p: retrofit_ceiling(retrofit_gem_by_province, p))

fig, axes = plt.subplots(1, 2, figsize=(17, 8))
draw_sensitivity_panel(axes[0], 'retrofit_ceiling_hazus', 'HAZUS: max achievable reduction via retrofit (04:17)', 'Blues',
                        eq_label_color='yellow',
                        eq_label_offsets={'Pazarcık Mw7.8': (8, -16), 'Elbistan Mw7.5': (-3, 8)},
                        legend_label='% reduction achievable (Factual -> full retrofit)')
draw_sensitivity_panel(axes[1], 'retrofit_ceiling_gem', 'GEM: max achievable reduction via retrofit (04:17)', 'Greens',
                        eq_label_offsets={'Pazarcık Mw7.8': (8, -16), 'Elbistan Mw7.5': (-3, 8)},
                        legend_label='% reduction achievable (Factual -> full retrofit)')
fig.suptitle('Retrofit ceiling by province: HAZUS vs GEM\n'
             '(max % reduction achievable, retrofitting from the 2023 Factual mix to 100% High-Code)', fontsize=13, y=0.98)
plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.savefig('map_retrofit_ceiling_hazus_vs_gem.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: map_retrofit_ceiling_hazus_vs_gem.png")

# ============================================================
# STEP 39: GEM 3x3 heatmap (national totals)
# ============================================================

df_matrix_gem = pd.DataFrame(
    {
        '04:17': [gem_matrix[('Pre-Code','04:17')], gem_matrix[('Low-Code','04:17')], gem_matrix[('High-Code','04:17')]],
        '14:00': [gem_matrix[('Pre-Code','14:00')], gem_matrix[('Low-Code','14:00')], gem_matrix[('High-Code','14:00')]],
        '17:00': [gem_matrix[('Pre-Code','17:00')], gem_matrix[('Low-Code','17:00')], gem_matrix[('High-Code','17:00')]],
    },
    index=['Pre-Code', 'Low-Code', 'High-Code']
)
df_matrix_gem.index.name = 'SQ2 (Compliance)'

fig, ax = plt.subplots(figsize=(6, 4.5))
matrix_vals_gem = df_matrix_gem.values
im = ax.imshow(matrix_vals_gem, cmap='Greens', aspect='auto')
ax.set_xticks(range(3))
ax.set_xticklabels(['04:17\n(night-time)', '14:00\n(day-time)', '17:00\n(transit)'])
ax.set_yticks(range(3))
ax.set_yticklabels(['Pre-Code', 'Low-Code', 'High-Code'])
for i in range(3):
    for j in range(3):
        val = matrix_vals_gem[i, j]
        text_color = 'white' if val > matrix_vals_gem.max() * 0.5 else 'black'
        label = f'{val:,.0f}' if val >= 1 else f'{val:.3f}'
        ax.text(j, i, label, ha='center', va='center', color=text_color, fontsize=11)
ax.set_title('GEM 3x3 matrix: national total deaths')
fig.colorbar(im, ax=ax, label='Deaths')
plt.tight_layout()
plt.savefig('gem_3x3_heatmap.png', dpi=150)
plt.close()
print("Saved: gem_3x3_heatmap.png")

# ============================================================
# STEP 40: Province-level heatmap, GEM version
# ============================================================

cell_order_gem = cell_order
province_list_gem = list(provinces.keys())
heat_data_gem = np.zeros((len(province_list_gem), len(cell_order_gem)))
for pi, prov in enumerate(province_list_gem):
    for ci, (era, time_label) in enumerate(cell_order_gem):
        occ_key = SQ1_OCC_KEYS[time_label]
        heat_data_gem[pi, ci] = compute_gem_deaths(provinces[prov], era, occ_key)

sort_idx_gem = np.argsort(-heat_data_gem[:, 0])
heat_data_gem_sorted = heat_data_gem[sort_idx_gem, :]
province_list_gem_sorted = [province_list_gem[i] for i in sort_idx_gem]

fig, ax = plt.subplots(figsize=(10, 7))
vmin_gem = max(heat_data_gem_sorted[heat_data_gem_sorted > 0].min(), 1e-3)
im = ax.imshow(heat_data_gem_sorted, cmap='Greens', aspect='auto',
               norm=LogNorm(vmin=vmin_gem, vmax=heat_data_gem_sorted.max()))
ax.set_xticks(range(len(cell_order_gem)))
ax.set_xticklabels([f'{e}\n{t}\n({TIME_DESC_ONLY[t]})' for e, t in cell_order_gem], fontsize=8)
ax.set_yticks(range(len(province_list_gem_sorted)))
ax.set_yticklabels(province_list_gem_sorted, fontsize=9)

for i in range(heat_data_gem_sorted.shape[0]):
    for j in range(heat_data_gem_sorted.shape[1]):
        val = heat_data_gem_sorted[i, j]
        rel = np.log(val + 1) / np.log(heat_data_gem_sorted.max() + 1)
        text_color = 'white' if rel > 0.55 else 'black'
        label = f'{val:,.0f}' if val >= 1 else f'{val:.3f}'
        ax.text(j, i, label, ha='center', va='center', color=text_color, fontsize=7)

ax.set_title('GEM: Deaths by province and scenario cell (log color scale, sorted by GEM Pre-Code x 04:17)')
fig.colorbar(im, ax=ax, label='Deaths (log scale)')
plt.tight_layout()
plt.savefig('province_level_heatmap_gem.png', dpi=150)
plt.close()
print("Saved: province_level_heatmap_gem.png")

# ============================================================
# STEP 41: SQ1 vs SQ2 sensitivity ratio chart -- GEM
# ============================================================

gem_sq1_ratios = {
    'Pre-Code':  gem_matrix[('Pre-Code','04:17')]  / gem_matrix[('Pre-Code','14:00')],
    'Low-Code':  gem_matrix[('Low-Code','04:17')]  / gem_matrix[('Low-Code','14:00')],
    'High-Code': gem_matrix[('High-Code','04:17')] / gem_matrix[('High-Code','14:00')],
}
gem_sq2_ratios = {
    '04:17': gem_matrix[('Pre-Code','04:17')] / gem_matrix[('High-Code','04:17')],
    '14:00': gem_matrix[('Pre-Code','14:00')] / gem_matrix[('High-Code','14:00')],
    '17:00': gem_matrix[('Pre-Code','17:00')] / gem_matrix[('High-Code','17:00')],
}

fig, ax = plt.subplots(figsize=(8, 5.5))
labels_gem = list(gem_sq1_ratios.keys()) + [TIME_LABELS_DISPLAY[t] for t in gem_sq2_ratios.keys()]
values_gem = list(gem_sq1_ratios.values()) + list(gem_sq2_ratios.values())
colors_gem = ['#2a78d6']*3 + ['#eda100']*3

bars = ax.bar(labels_gem, values_gem, color=colors_gem)
ax.set_ylabel('Ratio (worst / best)')
ax.set_title('SQ1 (timing) vs SQ2 (compliance) sensitivity ratio, GEM')
for bar, val in zip(bars, values_gem):
    ax.text(bar.get_x() + bar.get_width()/2, val + max(values_gem)*0.02, f'{val:.2f}', ha='center', fontsize=10)
plt.xticks(rotation=0, ha='center')

ax.annotate('SQ1 ratio, computed with compliance\nHELD FIXED at each level shown below',
            xy=(1, 0), xycoords=('data', 'axes fraction'),
            xytext=(0, -60), textcoords='offset points',
            ha='center', fontsize=7.5, color='#1a4e8a')
ax.annotate('SQ2 ratio, computed with timing\nHELD FIXED at each time shown below',
            xy=(4, 0), xycoords=('data', 'axes fraction'),
            xytext=(0, -60), textcoords='offset points',
            ha='center', fontsize=7.5, color='#a06800')

ax.set_ylim(0, max(values_gem) * 1.3)
ax.legend(handles=[
    Patch(color='#2a78d6', label='SQ1 ratio, per compliance level'),
    Patch(color='#eda100', label='SQ2 ratio, per timing'),
], loc='upper center', bbox_to_anchor=(0.5, -0.35), ncol=2, fontsize=8, frameon=False)

plt.subplots_adjust(bottom=0.38)
plt.savefig('sq1_sq2_sensitivity_gem.png', dpi=150, bbox_inches='tight')
plt.close()
print("Saved: sq1_sq2_sensitivity_gem.png")

# ============================================================
# STEP 42: Interaction effect analysis -- GEM version
# ============================================================

M_gem = np.array([
    [gem_matrix[('Pre-Code','04:17')],  gem_matrix[('Pre-Code','14:00')],  gem_matrix[('Pre-Code','17:00')]],
    [gem_matrix[('Low-Code','04:17')],  gem_matrix[('Low-Code','14:00')],  gem_matrix[('Low-Code','17:00')]],
    [gem_matrix[('High-Code','04:17')], gem_matrix[('High-Code','14:00')], gem_matrix[('High-Code','17:00')]],
])

logM_gem = np.log(M_gem)
grand_mean_gem = logM_gem.mean()
row_eff_gem = logM_gem.mean(axis=1) - grand_mean_gem
col_eff_gem = logM_gem.mean(axis=0) - grand_mean_gem

interaction_gem = np.zeros_like(logM_gem)
for i in range(3):
    for j in range(3):
        interaction_gem[i, j] = logM_gem[i, j] - grand_mean_gem - row_eff_gem[i] - col_eff_gem[j]

idx_gem = np.unravel_index(np.argmax(np.abs(interaction_gem)), interaction_gem.shape)
max_dev_pct_gem = (np.exp(interaction_gem[idx_gem]) - 1) * 100
print(f"\nLargest GEM interaction: {rows_lbl[idx_gem[0]]} x {cols_lbl[idx_gem[1]]} -- deviates {max_dev_pct_gem:+.1f}% from additive prediction")

interaction_list_gem = []
for i, r in enumerate(rows_lbl):
    for j, c in enumerate(cols_lbl):
        pct = (np.exp(interaction_gem[i, j]) - 1) * 100
        interaction_list_gem.append((f'{r} x {COLS_LBL_SHORT[c]}', pct))
interaction_list_gem.sort(key=lambda t: abs(t[1]))

fig, ax = plt.subplots(figsize=(7.5, 5))
labels_gem_int = [t[0] for t in interaction_list_gem]
vals_gem_int = [t[1] for t in interaction_list_gem]
colors_gem_int = ['#e34948' if v < 0 else '#2a78d6' for v in vals_gem_int]
ax.barh(labels_gem_int, vals_gem_int, color=colors_gem_int)
ax.axvline(0, color='black', linewidth=0.8)
ax.set_xlabel('Deviation from additive model (%)')
ax.set_title('GEM: Interaction residual by cell, ranked by |deviation|')
max_abs_gem = max(abs(v) for v in vals_gem_int)
ax.set_xlim(-max_abs_gem * 1.4, max_abs_gem * 1.4)
for i, v in enumerate(vals_gem_int):
    ax.text(v + (max_abs_gem*0.06 if v >= 0 else -max_abs_gem*0.06), i, f'{v:+.1f}%',
            va='center', ha='left' if v >= 0 else 'right', fontsize=9)
plt.tight_layout()
plt.savefig('tornado_interaction_residuals_gem.png', dpi=150)
plt.close()
print("Saved: tornado_interaction_residuals_gem.png")

print(f"\nHAZUS vs GEM: max interaction deviation comparison")
print(f"  HAZUS largest interaction: {hazus_max_interaction_cell[0]} x {hazus_max_interaction_cell[1]}  ({hazus_max_interaction_pct:+.1f}%)")
print(f"  GEM largest interaction:   {rows_lbl[idx_gem[0]]} x {cols_lbl[idx_gem[1]]}  ({max_dev_pct_gem:+.1f}%)")

# ============================================================
# STEP 43: Province-level SQ2 sensitivity tornado -- GEM version
# ============================================================

province_sq2_ratios_gem = []
for prov, d in provinces.items():
    deaths_pre_gem = compute_gem_deaths(d, 'Pre-Code', 'occ_night')
    deaths_high_gem = compute_gem_deaths(d, 'High-Code', 'occ_night')
    ratio_gem = deaths_high_gem / deaths_pre_gem if deaths_pre_gem > 0 else float('nan')
    province_sq2_ratios_gem.append((prov, ratio_gem, d['pga']))
province_sq2_ratios_gem.sort(key=lambda t: t[1])

fig, ax = plt.subplots(figsize=(7.5, 5.5))
labels_gem2 = [f"{t[0]} (PGA {t[2]:.2f})" for t in province_sq2_ratios_gem]
vals_gem2 = [t[1] for t in province_sq2_ratios_gem]
ax.barh(labels_gem2, vals_gem2, color='#3aa655')
ax.set_xlabel('High-Code deaths / Pre-Code deaths (lower = bigger benefit from compliance)')
ax.set_title('GEM: SQ2 sensitivity by province, ranked (at 04:17)')
ax.set_xlim(0, max(vals_gem2) * 1.15)
for i, v in enumerate(vals_gem2):
    ax.text(v + max(vals_gem2)*0.015, i, f'{v:.2f}', va='center', fontsize=8)
plt.tight_layout()
plt.savefig('tornado_province_sq2_sensitivity_gem.png', dpi=150)
plt.close()
print("Saved: tornado_province_sq2_sensitivity_gem.png")

hazus_order = [t[0] for t in province_sq2_ratios]
gem_order = [t[0] for t in province_sq2_ratios_gem]
print("\nRanking comparison (most-benefit-first), HAZUS vs GEM:")
print(f"{'Rank':6} {'HAZUS':15} {'GEM':15}")
for i in range(len(hazus_order)):
    print(f"{i+1:<6} {hazus_order[i]:15} {gem_order[i]:15}")

